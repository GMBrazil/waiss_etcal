from django.shortcuts import render, reverse, redirect
from django.http import HttpResponseRedirect, JsonResponse
from .models import Crop, Soil, Farm, Station, Data
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from .forms import NewUserForm
from decimal import Decimal
import openpyxl
import datetime

# Create your views here.

value_handler = ""
array_handler = list()

def index(request):
    return render(request, 'etcal/index.html')

def sign_up(request):
    user_form = NewUserForm(request.POST or None)
    if request.method == "POST":
        if user_form.is_valid():
            user = user_form.save()
            login(request, user)
            return render(request, 'etcal/index.html')
    context = {
        "user_form" : user_form
    }
    return render(request, 'registration/sign_up.html', context)


def database(request):
    return render(request, 'etcal/database.html')


def calculator(request):
    return render(request, 'etcal/calculator.html')


@login_required
def get_started(request):
    current_user = request.user
    farm_info = Farm.objects.filter(manager=current_user)
    crop_info = Crop.objects.all()
    soil_info = Soil.objects.all()
    station_info = Station.objects.all()

    global array_handler
    excel_data = list()

    if request.method == 'POST':
        #-----Farm Form-----#
        farm_name = request.POST["farm-name"]
        farm_prov = request.POST["farm-prov"]
        farm_muni = request.POST["farm-muni"]
        farm_brgy = request.POST["farm-brgy"]
        farm_lat = request.POST["farm-lat"]
        farm_lat_dir = request.POST["farm-lat-dir"]
        farm_long = request.POST["farm-long"]
        farm_long_dir = request.POST["farm-long-dir"]
        #-----Crop Form-----#
        crop_def = request.POST["crops"]
        crop_other = request.POST["crop-other"]
        crop_drz = request.POST["crop-drz"]
        crop_dtm = request.POST["crop-dtm"]
        stage_init = request.POST["stage-init"]
        stage_dev = request.POST["stage-dev"]
        stage_mid = request.POST["stage-mid"]
        stage_late = request.POST["stage-late"]
        kc_init = request.POST["kc-init"]
        kc_mid = request.POST["kc-mid"]
        kc_late = request.POST["kc-late"]
        #------Soil Form----#
        soil_def = request.POST["soils"]
        soil_other = request.POST["soil-other"]
        soil_fc = request.POST["soil-fc"]
        soil_pwp = request.POST["soil-pwp"]
        percolation = request.POST["soil-perc"]
        init_depl = request.POST["init-depl"]
        mad = request.POST["mad"]
        #------Station Form----#
        station_def = request.POST["stations"]
        station_other = request.POST["sta-name"]
        station_type = request.POST["sta-type"]
        station_elev = request.POST["sta-elev"]
        station_lat = request.POST["sta-lat"]
        station_lat_dir = request.POST["sta-lat-dir"]
        station_long = request.POST["sta-long"]
        station_long_dir = request.POST["sta-long-dir"]
        station_prov = request.POST["sta-prov"]
        station_muni = request.POST["sta-muni"]
        station_brgy = request.POST["sta-brgy"]
        #------Baseline Data----#
        dap = request.POST["dap"]
        corr_factor = request.POST["corr-factor"]
        date_measured = request.POST.getlist('date_measured[]')
        eto_data = request.POST.getlist('eto_data[]')
        rain_data = request.POST.getlist('rain_data[]')
        irrig_data = request.POST.getlist('irrig_data[]')

        # check if there are new inputs in crops, soils, stations
        if (crop_def == "") or (crop_def != "Other"):
            crop_type = crop_def
        elif (crop_other != "") and (crop_def == "Other"):
            crop_type = crop_other

        if (soil_def == "") or (soil_def != "Other"):
            soil_type = soil_def
        elif (soil_other != "") and (soil_def == "Other"):
            soil_type = soil_other

        if (station_def == "") or (station_def != "Other"):
            station_name = station_def
        elif (station_other != "") and (station_def == "Other"):
            station_name = station_other

        if (station_lat == "") or (station_long == ""):
            station_lat = None
            station_lat_dir = None
            station_long = None
            station_long_dir = None

        if (farm_lat == "") or (farm_long == ""):
            farm_lat = None
            farm_lat_dir = None
            farm_long = None
            farm_long_dir = None
        
        if (dap == "") or (corr_factor == ""):
            dap = None
            corr_factor = None

        # save to databases
        crop_data, created = Crop.objects.get_or_create(crop_type=crop_type, crop_drz=crop_drz, crop_dtm=crop_dtm, stage_init=stage_init,
                                                         stage_dev=stage_dev, stage_mid=stage_mid, stage_late=stage_late, kc_init=kc_init, kc_mid=kc_mid, kc_late=kc_late, author = current_user)
        crop_data.save()
        crop = Crop.objects.get(id=crop_data.id)

        soil_data, created = Soil.objects.get_or_create(
            soil_type=soil_type, soil_fc=soil_fc, soil_pwp=soil_pwp, percolation=percolation, init_depl=init_depl, mad=mad, author = current_user)
        soil_data.save()
        soil = Soil.objects.get(id=soil_data.id)

        station_data, created = Station.objects.get_or_create(station_name=station_name, station_type=station_type, station_elev=station_elev,
                                                              station_lat=station_lat, station_lat_dir=station_lat_dir, station_long=station_long, station_long_dir=station_long_dir, station_prov=station_prov, station_muni=station_muni, station_brgy=station_brgy)
        station_data.save()
        station = Station.objects.get(id=station_data.id)

        farm_data, created = Farm.objects.get_or_create(farm_name=farm_name, farm_prov=farm_prov, farm_muni=farm_muni, farm_brgy=farm_brgy,
                                                        farm_lat=farm_lat, farm_lat_dir=farm_lat_dir, farm_long=farm_long, farm_long_dir=farm_long_dir, date_planted=dap, crop=crop, soil=soil, corr_factor=corr_factor, manager=current_user)
        farm_data.save()
        farm = Farm.objects.get(id=farm_data.id)

        #save farm id to global to access in dash function after redirecting, for organizing farms in the select tab in dash
        global value_handler
        value_handler = farm
        
        excel_data = array_handler
        if excel_data:
            for row in excel_data[1:]:
                date = datetime.datetime.strptime(row[0],"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
                eto = row[1]
                rainfall = row[2]
                irrigation = row[3]
                data, created = Data.objects.get_or_create(farm=farm, station=station, timestamp=date, eto=eto, rainfall=rainfall, irrigation=irrigation)
                data.save()
            return HttpResponseRedirect(reverse('etcal:dashboard'))

        elif (date_measured != "") or (eto_data != "") or (rain_data != "") or (irrig_data != ""):
            for date, eto, rainfall, irrigation in zip(date_measured, eto_data, rain_data, irrig_data):
                if not date:
                    continue
                elif (date != "") and (eto != "") and ((rainfall == "") or (irrigation == "")):
                    rainfall = 0
                    irrigation = 0
                else:
                    eto = Decimal(eto)
                    rainfall = Decimal(rainfall)
                    irrigation = Decimal(irrigation)
                data, created = Data.objects.get_or_create(farm=farm, station=station, timestamp=date, eto=eto, rainfall=rainfall, irrigation=irrigation)
                data.save()
            return HttpResponseRedirect(reverse('etcal:dashboard'))


    context = {
        "crop_info": crop_info,
        "soil_info": soil_info,
        "station_info": station_info,
        "excel_data": excel_data
    }

    excel_data = list()
    array_handler = list()

    return render(request, 'etcal/get-started.html', context)

@login_required
def get_started_upload_file(request):
    current_user = request.user
    farm_info = Farm.objects.filter(manager=current_user)
    crop_info = Crop.objects.all()
    soil_info = Soil.objects.all()
    station_info = Station.objects.all()

    context = {
        "crop_info": crop_info,
        "soil_info": soil_info,
        "station_info": station_info
    }

    if "GET" == request.method:
        return render(request, 'etcal/upload-file.html', context)
    
    else:
        excel_file = request.FILES["excel_file"]
        workbook = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = workbook["Sheet1"]
        print(worksheet)
        excel_data = list()
        # iterating over the rows and getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        global array_handler
        array_handler = excel_data

        context = {
            "excel_data": excel_data,
            "message": "File has been uploaded! Your data are shown below."
        }

        return render(request, 'etcal/upload-file.html', context)

@login_required
def dashboard(request):
    current_user = request.user
    farm_info = Farm.objects.filter(manager=current_user)
    if not farm_info:
        return HttpResponseRedirect(reverse('etcal:get_started'))
    crop_info = Crop.objects.all()
    soil_info = Soil.objects.all()
    station_info = Station.objects.all()
    
    global value_handler
    new_farm = value_handler

    context = {
        "new_farm" : new_farm,
        "farm_info": farm_info,
        "crop_info": crop_info,
        "soil_info": soil_info,
        "station_info": station_info
    }

    new_farm = ""
    value_handler = ""

    return render(request, 'etcal/dashboard.html', context)


def contribute(request):
    return render(request, 'etcal/contribute.html')


def load_crop(request):
    crop_id = request.GET.get('crop')
    crops = Crop.objects.filter(crop_type=crop_id)

    context = {
        "crops": crops,
    }
    return render(request, 'etcal/load-crop.html', context)


def load_soil(request):
    soil_id = request.GET.get('soil')
    soils = Soil.objects.filter(soil_type=soil_id)

    context = {
        "soils": soils,
    }
    return render(request, 'etcal/load-soil.html', context)


def load_station(request):
    station_id = request.GET.get('station')
    stations = Station.objects.filter(station_name=station_id)

    context = {
        "stations": stations,
    }
    return render(request, 'etcal/load-station.html', context)


def load_dash(request):
	farm_id = request.GET.get('farm')
	farm_details = Farm.objects.get(id=farm_id)
	crop_data = Crop.objects.get(crop_type=farm_details.crop)
	soil_data = Soil.objects.get(soil_type=farm_details.soil)
	latest_farm_data = Data.objects.filter(farm=farm_details).order_by('timestamp')

	context = {
		"farm_details": farm_details,
		"crop_data": crop_data,
		"soil_data": soil_data,
		"latest_farm_data": latest_farm_data,
	}
	return render(request, 'etcal/load-dash.html', context)

def load_file(request):
    current_user = request.user
    farm_info = Farm.objects.filter(manager=current_user)
    crop_info = Crop.objects.all()
    soil_info = Soil.objects.all()
    station_info = Station.objects.all()

    global array_handler
    array_handler = list()
    excel_data = list()

    context = {
        "crop_info": crop_info,
        "soil_info": soil_info,
        "station_info": station_info
    }

    if "GET" == request.method:
        return render(request, 'etcal/get-started.html', context)
    
    else:
        excel_file = request.FILES["excel_file"]
        workbook = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = workbook["Sheet1"]
        print(worksheet)
        # iterating over the rows and getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        array_handler = excel_data

        context = {
            "excel_data": excel_data
        }

        return render(request, 'etcal/load-file.html', context)